#!/usr/bin/env python3
"""
CLI tool to build a multi-company comp sheet Excel file from a context JSON.

Usage:
    python infra/comp_builder.py --context context.json --output comp_sheet.xlsx

Produces an 8-tab workbook:
  1. Comp Summary — one-pager comp table with implied valuation
  2. Revenue Drivers — unit economics decomposition per company
  3. Operating KPIs — broader operational metrics comparison
  4. Financial Summary — side-by-side income statements
  5. Growth & Margins — trend analysis across companies
  6. Valuation Detail — expanded valuation with sensitivities
  7. Balance Sheet & Capital — leverage and capital returns
  8. Raw Data — full quarterly appendix per company
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants (matching excel_builder.py design system)
# ---------------------------------------------------------------------------

NAVY_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
TARGET_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
MEDIAN_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
WHITE_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
THIN_SIDE = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BOLD_SIDE = Side(style="thin", color="000000")
BOLD_BORDER = Border(left=BOLD_SIDE, right=BOLD_SIDE, top=BOLD_SIDE, bottom=BOLD_SIDE)
HEADER_BORDER = Border(bottom=Side(style="medium", color="1B2A4A"))

GREEN_FONT = Font(name="Calibri", size=10, color="006100")
RED_FONT = Font(name="Calibri", size=10, color="9C0006")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

COL_MIN_WIDTH = 13
COL_MAX_WIDTH = 22

_STYLES_REGISTERED = False


def register_styles(wb: Workbook) -> None:
    global _STYLES_REGISTERED
    if _STYLES_REGISTERED:
        return
    _STYLES_REGISTERED = True

    header = NamedStyle(name="header")
    header.font = WHITE_FONT
    header.fill = NAVY_FILL
    header.border = HEADER_BORDER
    header.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(header)

    data = NamedStyle(name="data")
    data.font = Font(name="Calibri", size=10)
    data.border = THIN_BORDER
    wb.add_named_style(data)

    data_pct = NamedStyle(name="data_pct")
    data_pct.font = Font(name="Calibri", size=10)
    data_pct.number_format = "0.0%"
    data_pct.border = THIN_BORDER
    wb.add_named_style(data_pct)

    data_currency = NamedStyle(name="data_currency")
    data_currency.font = Font(name="Calibri", size=10)
    data_currency.number_format = "#,##0"
    data_currency.border = THIN_BORDER
    wb.add_named_style(data_currency)

    input_cell = NamedStyle(name="input_cell")
    input_cell.font = Font(name="Calibri", size=10, bold=True)
    input_cell.fill = TARGET_FILL
    input_cell.border = BOLD_BORDER
    wb.add_named_style(input_cell)


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _col(col: int) -> str:
    return get_column_letter(col)


def _auto_col_widths(ws, min_w: int = COL_MIN_WIDTH, max_w: int = COL_MAX_WIDTH) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = _col(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        width = min(max(max_len + 2, min_w), max_w)
        ws.column_dimensions[col_letter].width = width


def _freeze(ws, row: int = 2, col: int = 2) -> None:
    ws.freeze_panes = ws.cell(row=row, column=col)


def _header_row(ws, row: int, values: list, start_col: int = 1) -> None:
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.style = "header"


def _is_pct(name: str) -> bool:
    lower = name.lower()
    return any(kw in lower for kw in [
        "margin", "growth", "yield", "rate", "%", "pct", "percent", "retention"
    ])


def _is_multiple(name: str) -> bool:
    lower = name.lower()
    return any(kw in lower for kw in [
        "p/e", "pe", "ev/ebitda", "p/s", "p/b", "ev/fcf", "peg"
    ])


def _fmt(name: str) -> str:
    if _is_pct(name):
        return "0.0%"
    if _is_multiple(name):
        return "0.0x"
    return "#,##0"


def _data_cell(ws, row: int, col: int, value, name: str = "",
               is_target: bool = False, is_median: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.style = "data"
    cell.number_format = _fmt(name)
    if is_target:
        cell.fill = TARGET_FILL
        cell.font = Font(name="Calibri", size=10, bold=True)
    elif is_median:
        cell.fill = MEDIAN_FILL
        cell.font = Font(name="Calibri", size=10, bold=True)
    return cell


def _add_cond_fmt(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    if min_col > max_col or min_row > max_row:
        return
    rng = f"{_col(min_col)}{min_row}:{_col(max_col)}{max_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThan", formula=["0"], font=GREEN_FONT))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT))


def _companies(ctx: dict) -> list[dict]:
    return ctx.get("companies", [])


def _target(ctx: dict) -> dict | None:
    for c in _companies(ctx):
        if c.get("is_target"):
            return c
    return _companies(ctx)[0] if _companies(ctx) else None


# ---------------------------------------------------------------------------
# Tab 1: Comp Summary
# ---------------------------------------------------------------------------

def build_comp_summary(wb: Workbook, ctx: dict) -> None:
    """Transposed layout: metrics in Column A, companies across the top."""
    companies = _companies(ctx)
    if not companies:
        return

    ws = wb.create_sheet("Comp Summary")

    # Sort: target first, then peers by market cap descending
    sorted_cos = sorted(companies, key=lambda c: (
        0 if c.get("is_target") else 1,
        -(c.get("market_data", {}).get("market_cap") or 0)
    ))

    n_cos = len(sorted_cos)
    peers = [c for c in sorted_cos if not c.get("is_target")]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cos + 3)
    title = ws.cell(row=1, column=1,
                    value=f"Comparable Companies Analysis — {ctx.get('target_ticker', '')} vs Peers")
    title.font = Font(name="Calibri", size=14, bold=True, color="1B2A4A")
    ws.cell(row=2, column=1, value=f"As of {ctx.get('as_of_date', 'N/A')}").font = Font(
        name="Calibri", size=9, italic=True, color="666666")

    # Row 4: header — Metric | Company1 | Company2 | ... | Peer Median | Peer Mean
    header_vals = ["Metric"] + [co.get("ticker", "") for co in sorted_cos]
    if len(peers) >= 2:
        header_vals += ["Peer Median", "Peer Mean"]
    _header_row(ws, 4, header_vals)

    # Map column index for each company (1-indexed, col 1 = metric label)
    co_cols = {i: i + 2 for i in range(n_cos)}  # company i → column i+2
    median_col = n_cos + 2 if len(peers) >= 2 else None
    mean_col = n_cos + 3 if len(peers) >= 2 else None

    # Identify which columns are peers (for MEDIAN/AVERAGE formulas)
    peer_col_letters = []
    for i, co in enumerate(sorted_cos):
        if not co.get("is_target"):
            peer_col_letters.append(_col(co_cols[i]))

    # Metric rows
    metrics = [
        ("Price", "price", "$#,##0.00"),
        ("Market Cap ($mm)", "market_cap", "#,##0"),
        ("Enterprise Value ($mm)", "enterprise_value", "#,##0"),
        ("Trailing P/E", "trailing_pe", "0.0x"),
        ("Forward P/E", "forward_pe", "0.0x"),
        ("EV/EBITDA", "ev_ebitda", "0.0x"),
        ("P/S", "price_to_sales", "0.0x"),
        ("EV/FCF", "ev_fcf", "0.0x"),
        ("Dividend Yield", "dividend_yield", "0.0%"),
        ("Rev Growth (Latest Q)", "rev_growth", "0.0%"),
        ("Op Margin (Latest Q)", "op_margin", "0.0%"),
        ("Net Margin (Latest Q)", "net_margin", "0.0%"),
        ("FCF Margin (Latest Q)", "fcf_margin", "0.0%"),
        ("Beta", "beta", "0.00"),
    ]

    row = 5
    for label, key, fmt in metrics:
        # Metric label in column A
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.style = "data"
        label_cell.font = Font(name="Calibri", size=10, bold=True)

        # Company values
        for i, co in enumerate(sorted_cos):
            md = co.get("market_data", {})
            growth = co.get("growth", {})
            margins = co.get("margins", {})
            periods = co.get("periods", [])
            latest_p = periods[-1] if periods else None
            is_tgt = co.get("is_target", False)

            if key == "rev_growth":
                val = _latest_val(growth, "Revenue Growth YoY", latest_p)
            elif key == "op_margin":
                val = _latest_val(margins, "Operating Margin", latest_p)
            elif key == "net_margin":
                val = _latest_val(margins, "Net Margin", latest_p)
            elif key == "fcf_margin":
                val = _latest_val(margins, "FCF Margin", latest_p)
            else:
                val = md.get(key)

            cell = ws.cell(row=row, column=co_cols[i], value=val)
            cell.style = "data"
            cell.number_format = fmt
            if is_tgt:
                cell.fill = TARGET_FILL
                cell.font = Font(name="Calibri", size=10, bold=True)

        # Peer Median / Mean formulas
        if median_col and peer_col_letters:
            refs = ",".join(f"{cl}{row}" for cl in peer_col_letters)
            med_cell = ws.cell(row=row, column=median_col,
                               value=f"=MEDIAN({refs})")
            med_cell.style = "data"
            med_cell.fill = MEDIAN_FILL
            med_cell.font = Font(name="Calibri", size=10, bold=True)
            med_cell.number_format = fmt

            avg_cell = ws.cell(row=row, column=mean_col,
                               value=f"=AVERAGE({refs})")
            avg_cell.style = "data"
            avg_cell.font = Font(name="Calibri", size=10, bold=True)
            avg_cell.number_format = fmt

        row += 1

    # Implied valuation section
    implied = ctx.get("implied_valuation", {})
    if implied:
        row += 1
        ws.cell(row=row, column=1, value="Implied Valuation").font = Font(
            name="Calibri", size=11, bold=True, color="1B2A4A")
        row += 1

        target_col = None
        for i, co in enumerate(sorted_cos):
            if co.get("is_target"):
                target_col = co_cols[i]
                break

        impl_metrics = [
            ("P/E Implied Price", implied.get("pe_implied")),
            ("EV/EBITDA Implied Price", implied.get("ev_ebitda_implied")),
            ("P/S Implied Price", implied.get("ps_implied")),
            ("EV/FCF Implied Price", implied.get("ev_fcf_implied")),
            ("Median Implied Price", implied.get("median_implied")),
        ]
        for imp_label, val in impl_metrics:
            if val is None:
                continue
            ws.cell(row=row, column=1, value=imp_label).style = "data"
            ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
            if target_col:
                cell = ws.cell(row=row, column=target_col, value=val)
                cell.style = "data"
                cell.number_format = "$#,##0.00"
                cell.fill = MEDIAN_FILL
                cell.font = Font(name="Calibri", size=10, bold=True)
            row += 1

    _freeze(ws, row=5, col=2)
    _auto_col_widths(ws)


def _latest_val(data: dict, key: str, period: str | None):
    if not data or not period:
        return None
    series = data.get(key, {})
    return series.get(period)


# ---------------------------------------------------------------------------
# Tab 2: Revenue Drivers
# ---------------------------------------------------------------------------

def build_revenue_drivers(wb: Workbook, ctx: dict) -> None:
    companies = _companies(ctx)
    has_kpis = any(c.get("kpis") for c in companies)
    if not has_kpis:
        return

    ws = wb.create_sheet("Revenue Drivers")
    ws.cell(row=1, column=1,
            value="Revenue Driver Decomposition").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A")
    ws.cell(row=2, column=1,
            value="Unit economics by company — trailing 4 quarters").font = Font(
        name="Calibri", size=9, italic=True, color="666666")

    current_row = 4

    for co in companies:
        kpis = co.get("kpis", {})
        categories = co.get("kpi_categories", {})
        periods = co.get("periods", [])
        # Use last 4 periods for revenue drivers
        display_periods = periods[-4:] if len(periods) >= 4 else periods
        if not kpis or not display_periods:
            continue

        is_tgt = co.get("is_target", False)

        # Company section header
        label = f"{co.get('name', '')} ({co.get('ticker', '')})"
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="1B2A4A")
        if is_tgt:
            cell.fill = TARGET_FILL
        current_row += 1

        _header_row(ws, current_row, ["Metric"] + display_periods)
        current_row += 1

        # Group by category if available, else flat
        seg_kpis = categories.get("Segment Revenue", [])
        growth_kpis = categories.get("Growth KPIs", [])
        other_kpis = [k for k in kpis if k not in seg_kpis and k not in growth_kpis]
        ordered = seg_kpis + growth_kpis + other_kpis

        for kpi_name in ordered:
            series = kpis.get(kpi_name, {})
            if not series:
                continue

            label_cell = ws.cell(row=current_row, column=1, value=kpi_name)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10,
                                   bold=(kpi_name in seg_kpis))
            if is_tgt:
                label_cell.fill = TARGET_FILL

            for j, p in enumerate(display_periods):
                val = series.get(p)
                c = ws.cell(row=current_row, column=j + 2, value=val)
                c.style = "data"
                c.number_format = _fmt(kpi_name)
                if is_tgt:
                    c.fill = TARGET_FILL

            current_row += 1

        current_row += 1  # blank separator between companies

    _freeze(ws, row=5, col=2)
    _auto_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 3: Operating KPIs
# ---------------------------------------------------------------------------

def build_operating_kpis(wb: Workbook, ctx: dict) -> None:
    companies = _companies(ctx)
    has_kpis = any(c.get("kpis") for c in companies)
    if not has_kpis:
        return

    ws = wb.create_sheet("Operating KPIs")
    ws.cell(row=1, column=1,
            value="Operating KPI Comparison").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A")
    ws.cell(row=2, column=1,
            value="Latest available quarter per company").font = Font(
        name="Calibri", size=9, italic=True, color="666666")

    # Build a union of all efficiency/engagement KPIs across companies
    efficiency_kpis: set[str] = set()
    for co in companies:
        cats = co.get("kpi_categories", {})
        for cat_name, kpi_list in cats.items():
            if cat_name != "Segment Revenue":
                efficiency_kpis.update(kpi_list)
        # Also add any kpis not categorized
        for k in co.get("kpis", {}):
            efficiency_kpis.add(k)

    if not efficiency_kpis:
        return

    kpi_names = sorted(efficiency_kpis)

    # Header: KPI name | Company1 | Company2 | ...
    tickers = [co.get("ticker", "") for co in companies]
    _header_row(ws, 4, ["KPI"] + tickers)

    row = 5
    for kpi_name in kpi_names:
        ws.cell(row=row, column=1, value=kpi_name).style = "data"
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)

        for j, co in enumerate(companies):
            kpis = co.get("kpis", {})
            periods = co.get("periods", [])
            series = kpis.get(kpi_name, {})
            # Get latest value
            val = None
            if series and periods:
                for p in reversed(periods):
                    if p in series and series[p] is not None:
                        val = series[p]
                        break

            cell = ws.cell(row=row, column=j + 2, value=val)
            cell.style = "data"
            cell.number_format = _fmt(kpi_name)
            if co.get("is_target"):
                cell.fill = TARGET_FILL

        row += 1

    # Footnote
    row += 1
    ws.cell(row=row, column=1,
            value="Note: Blank cells indicate KPI not reported by that company.").font = Font(
        name="Calibri", size=9, italic=True, color="666666")

    _freeze(ws, row=5, col=2)
    _auto_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 4: Financial Summary
# ---------------------------------------------------------------------------

def build_financial_summary(wb: Workbook, ctx: dict) -> None:
    companies = _companies(ctx)
    if not companies:
        return

    ws = wb.create_sheet("Financial Summary")
    ws.cell(row=1, column=1,
            value="Financial Summary — Side-by-Side").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A")

    fin_metrics = [
        "Revenue", "Gross Profit", "Operating Income", "EBITDA",
        "Net Income", "EPS",
    ]

    current_row = 3

    for co in companies:
        financials = co.get("financials", {})
        periods = co.get("periods", [])
        display_periods = periods[-4:] if len(periods) >= 4 else periods
        if not financials or not display_periods:
            continue

        is_tgt = co.get("is_target", False)

        # Company header
        label = f"{co.get('name', '')} ({co.get('ticker', '')})"
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="1B2A4A")
        if is_tgt:
            cell.fill = TARGET_FILL
        current_row += 1

        _header_row(ws, current_row, ["Metric"] + display_periods)
        current_row += 1

        for metric in fin_metrics:
            series = financials.get(metric, {})
            if not series:
                continue

            label_cell = ws.cell(row=current_row, column=1, value=metric)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True)
            if is_tgt:
                label_cell.fill = TARGET_FILL

            for j, p in enumerate(display_periods):
                val = series.get(p)
                c = ws.cell(row=current_row, column=j + 2, value=val)
                c.style = "data"
                c.number_format = "#,##0.00" if metric == "EPS" else "#,##0"
                if is_tgt:
                    c.fill = TARGET_FILL

            current_row += 1

        # Add margins
        margins = co.get("margins", {})
        for margin_name in ["Gross Margin", "Operating Margin", "Net Margin"]:
            series = margins.get(margin_name, {})
            if not series:
                continue
            label_cell = ws.cell(row=current_row, column=1, value=margin_name)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, italic=True)

            for j, p in enumerate(display_periods):
                val = series.get(p)
                c = ws.cell(row=current_row, column=j + 2, value=val)
                c.style = "data_pct"
                if is_tgt:
                    c.fill = TARGET_FILL

            current_row += 1

        current_row += 1  # separator

    _freeze(ws)
    _auto_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 5: Growth & Margins
# ---------------------------------------------------------------------------

def build_growth_margins(wb: Workbook, ctx: dict) -> None:
    companies = _companies(ctx)
    if not companies:
        return

    ws = wb.create_sheet("Growth & Margins")
    ws.cell(row=1, column=1,
            value="Growth & Margin Trends").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A")

    trend_metrics = [
        ("Revenue Growth YoY", "growth"),
        ("EPS Growth YoY", "growth"),
        ("Gross Margin", "margins"),
        ("Operating Margin", "margins"),
        ("FCF Margin", "margins"),
    ]

    current_row = 3

    for metric_name, source_key in trend_metrics:
        ws.cell(row=current_row, column=1, value=metric_name).font = Font(
            name="Calibri", size=11, bold=True, color="1B2A4A")
        current_row += 1

        # Get union of periods across companies for this metric
        all_periods: list[str] = []
        for co in companies:
            data = co.get(source_key, {})
            series = data.get(metric_name, {})
            for p in co.get("periods", []):
                if p in series and p not in all_periods:
                    all_periods.append(p)

        display_periods = all_periods[-8:] if len(all_periods) >= 8 else all_periods
        if not display_periods:
            current_row += 1
            continue

        _header_row(ws, current_row, ["Company"] + display_periods)
        current_row += 1

        for co in companies:
            is_tgt = co.get("is_target", False)
            data = co.get(source_key, {})
            series = data.get(metric_name, {})

            label_cell = ws.cell(row=current_row, column=1,
                                 value=co.get("ticker", ""))
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True)
            if is_tgt:
                label_cell.fill = TARGET_FILL

            for j, p in enumerate(display_periods):
                val = series.get(p)
                c = ws.cell(row=current_row, column=j + 2, value=val)
                c.style = "data_pct"
                if is_tgt:
                    c.fill = TARGET_FILL

            current_row += 1

        # Conditional formatting on data area
        data_start = current_row - len(companies)
        data_end = current_row - 1
        if len(display_periods) > 0:
            _add_cond_fmt(ws, data_start, data_end, 2, len(display_periods) + 1)

        current_row += 1  # separator

    _freeze(ws)
    _auto_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 6: Valuation Detail
# ---------------------------------------------------------------------------

def build_valuation_detail(wb: Workbook, ctx: dict) -> None:
    companies = _companies(ctx)
    target = _target(ctx)
    implied = ctx.get("implied_valuation", {})
    if not target or not implied:
        return

    ws = wb.create_sheet("Valuation Detail")
    ws.cell(row=1, column=1,
            value=f"Valuation Detail — {target.get('ticker', '')}").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A")

    # Implied prices from each methodology
    row = 3
    ws.cell(row=row, column=1, value="Implied Share Price by Methodology").font = Font(
        name="Calibri", size=11, bold=True, color="1B2A4A")
    row += 1

    current_price = target.get("market_data", {}).get("price")

    _header_row(ws, row, ["Methodology", "Peer Median Multiple",
                           "Target Metric", "Implied Price", "vs Current"])
    row += 1

    methodologies = [
        ("P/E", "pe_implied"),
        ("EV/EBITDA", "ev_ebitda_implied"),
        ("P/S", "ps_implied"),
        ("EV/FCF", "ev_fcf_implied"),
    ]

    for method_name, key in methodologies:
        implied_price = implied.get(key)
        if implied_price is None:
            continue

        ws.cell(row=row, column=1, value=method_name).style = "data"
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)

        # Implied price
        cell = ws.cell(row=row, column=4, value=implied_price)
        cell.style = "data"
        cell.number_format = "$#,##0.00"

        # vs current
        if current_price and current_price > 0:
            upside = (implied_price - current_price) / current_price
            up_cell = ws.cell(row=row, column=5, value=upside)
            up_cell.style = "data_pct"
            up_cell.font = GREEN_FONT if upside > 0 else RED_FONT

        row += 1

    # Median implied
    median_implied = implied.get("median_implied")
    if median_implied:
        row += 1
        ws.cell(row=row, column=1, value="Median Implied Price").style = "data"
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=1).fill = MEDIAN_FILL
        cell = ws.cell(row=row, column=4, value=median_implied)
        cell.style = "data"
        cell.number_format = "$#,##0.00"
        cell.fill = MEDIAN_FILL
        if current_price and current_price > 0:
            upside = (median_implied - current_price) / current_price
            up_cell = ws.cell(row=row, column=5, value=upside)
            up_cell.style = "data_pct"
            up_cell.font = GREEN_FONT if upside > 0 else RED_FONT
            up_cell.fill = MEDIAN_FILL

    # Premium/discount section
    row += 3
    ws.cell(row=row, column=1, value="Premium / Discount vs Peers").font = Font(
        name="Calibri", size=11, bold=True, color="1B2A4A")
    row += 1

    mult_fields = [
        ("P/E", "trailing_pe"),
        ("EV/EBITDA", "ev_ebitda"),
        ("P/S", "price_to_sales"),
        ("EV/FCF", "ev_fcf"),
    ]

    _header_row(ws, row, ["Multiple", "Target", "Peer Median", "Premium/Discount"])
    row += 1

    target_md = target.get("market_data", {})
    peers = [c for c in companies if not c.get("is_target")]

    for mult_name, key in mult_fields:
        tgt_val = target_md.get(key)
        peer_vals = [c.get("market_data", {}).get(key) for c in peers
                     if c.get("market_data", {}).get(key) is not None]

        ws.cell(row=row, column=1, value=mult_name).style = "data"
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)

        if tgt_val is not None:
            ws.cell(row=row, column=2, value=tgt_val).style = "data"
            ws.cell(row=row, column=2).number_format = "0.0x"

        if peer_vals:
            peer_vals_sorted = sorted(peer_vals)
            median_val = peer_vals_sorted[len(peer_vals_sorted) // 2]
            ws.cell(row=row, column=3, value=median_val).style = "data"
            ws.cell(row=row, column=3).number_format = "0.0x"

            if tgt_val is not None and median_val and median_val > 0:
                prem = (tgt_val - median_val) / median_val
                prem_cell = ws.cell(row=row, column=4, value=prem)
                prem_cell.style = "data_pct"

        row += 1

    _auto_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 7: Balance Sheet & Capital
# ---------------------------------------------------------------------------

def build_balance_sheet_capital(wb: Workbook, ctx: dict) -> None:
    companies = _companies(ctx)
    if not companies:
        return

    # Check if any company has financials with BS/capital metrics
    bs_metrics = [
        "Net Debt", "Net Debt/EBITDA", "Interest Coverage",
        "FCF Yield", "Shareholder Yield", "Buyback Yield", "Dividend Yield",
    ]

    ws = wb.create_sheet("Balance Sheet & Capital")
    ws.cell(row=1, column=1,
            value="Leverage & Capital Returns").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A")

    tickers = [co.get("ticker", "") for co in companies]
    _header_row(ws, 3, ["Metric"] + tickers)

    row = 4
    for metric in bs_metrics:
        ws.cell(row=row, column=1, value=metric).style = "data"
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)

        for j, co in enumerate(companies):
            # Try market_data first, then financials
            md = co.get("market_data", {})
            financials = co.get("financials", {})
            margins = co.get("margins", {})
            periods = co.get("periods", [])
            latest_p = periods[-1] if periods else None

            val = md.get(metric.lower().replace(" ", "_").replace("/", "_"))
            if val is None and latest_p:
                series = financials.get(metric, margins.get(metric, {}))
                val = series.get(latest_p) if isinstance(series, dict) else None

            cell = ws.cell(row=row, column=j + 2, value=val)
            cell.style = "data"
            cell.number_format = _fmt(metric)
            if co.get("is_target"):
                cell.fill = TARGET_FILL

        row += 1

    _freeze(ws, row=4, col=2)
    _auto_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 8: Raw Data
# ---------------------------------------------------------------------------

def build_raw_data(wb: Workbook, ctx: dict) -> None:
    companies = _companies(ctx)
    if not companies:
        return

    ws = wb.create_sheet("Raw Data")
    ws.cell(row=1, column=1,
            value="Raw Quarterly Data — All Companies").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A")

    current_row = 3

    for co in companies:
        financials = co.get("financials", {})
        kpis = co.get("kpis", {})
        margins = co.get("margins", {})
        growth = co.get("growth", {})
        periods = co.get("periods", [])
        is_tgt = co.get("is_target", False)

        if not periods:
            continue

        # Company header
        label = f"{co.get('name', '')} ({co.get('ticker', '')})"
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="1B2A4A")
        if is_tgt:
            cell.fill = TARGET_FILL
        current_row += 1

        _header_row(ws, current_row, ["Metric"] + periods)
        current_row += 1

        # All data sources combined
        all_series = {}
        all_series.update(financials)
        all_series.update(margins)
        all_series.update(growth)
        all_series.update(kpis)

        for metric, series in all_series.items():
            if not isinstance(series, dict):
                continue

            label_cell = ws.cell(row=current_row, column=1, value=metric)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10,
                                   bold=not _is_pct(metric))

            for j, p in enumerate(periods):
                val = series.get(p)
                c = ws.cell(row=current_row, column=j + 2, value=val)
                c.style = "data"
                c.number_format = _fmt(metric)

            current_row += 1

        current_row += 2  # separator between companies

    _freeze(ws)
    _auto_col_widths(ws)


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

TAB_BUILDERS = [
    build_comp_summary,
    build_revenue_drivers,
    build_operating_kpis,
    build_financial_summary,
    build_growth_margins,
    build_valuation_detail,
    build_balance_sheet_capital,
    build_raw_data,
]


def build_workbook(ctx: dict) -> Workbook:
    global _STYLES_REGISTERED
    _STYLES_REGISTERED = False

    wb = Workbook()
    register_styles(wb)

    default_sheet = wb.active
    tabs_created = 0

    for builder_fn in TAB_BUILDERS:
        try:
            builder_fn(wb, ctx)
            tabs_created += 1
        except Exception as e:
            print(f"Warning: {builder_fn.__name__} failed: {e}", file=sys.stderr)

    if tabs_created > 0 and default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a multi-company comp sheet Excel file from context JSON."
    )
    parser.add_argument(
        "--context", required=True,
        help="Path to the input context JSON file."
    )
    parser.add_argument(
        "--output", required=True,
        help="Path for the output .xlsx file."
    )
    args = parser.parse_args()

    context_path = Path(args.context)
    if not context_path.is_file():
        print(f"Error: context file not found: {args.context}", file=sys.stderr)
        sys.exit(1)

    try:
        with open(context_path, "r") as f:
            ctx = json.load(f)
    except json.JSONDecodeError as e:
        print(f"Error: invalid JSON in context file: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        wb = build_workbook(ctx)
    except Exception as e:
        print(f"Error building workbook: {e}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(output_path))
    except Exception as e:
        print(f"Error saving workbook: {e}", file=sys.stderr)
        sys.exit(1)

    print(str(output_path))


if __name__ == "__main__":
    main()
