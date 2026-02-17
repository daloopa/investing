#!/usr/bin/env python3
"""
CLI tool to build a multi-tab financial model Excel file from a context JSON.

Usage:
    python infra/excel_builder.py --context context.json --output model.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

_STYLES_REGISTERED = False

NAVY_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
PROJECTED_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WHITE_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
THIN_SIDE = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BOLD_SIDE = Side(style="thin", color="000000")
BOLD_BORDER = Border(left=BOLD_SIDE, right=BOLD_SIDE, top=BOLD_SIDE, bottom=BOLD_SIDE)
HEADER_BORDER = Border(bottom=Side(style="medium", color="1B2A4A"))

GREEN_FONT = Font(name="Calibri", size=10, color="006100")
RED_FONT = Font(name="Calibri", size=10, color="9C0006")

COL_MIN_WIDTH = 12
COL_MAX_WIDTH = 20


def register_styles(wb: Workbook) -> None:
    """Register reusable NamedStyles on the workbook (idempotent)."""
    global _STYLES_REGISTERED
    if _STYLES_REGISTERED:
        return
    _STYLES_REGISTERED = True

    header = NamedStyle(name="header")
    header.font = WHITE_FONT
    header.fill = NAVY_FILL
    header.border = HEADER_BORDER
    header.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(header)

    data = NamedStyle(name="data")
    data.font = Font(name="Calibri", size=10)
    data.border = THIN_BORDER
    wb.add_named_style(data)

    data_pct = NamedStyle(name="data_pct")
    data_pct.font = Font(name="Calibri", size=10)
    data_pct.number_format = "0.0%"
    data_pct.border = THIN_BORDER
    wb.add_named_style(data_pct)

    data_currency = NamedStyle(name="data_currency")
    data_currency.font = Font(name="Calibri", size=10)
    data_currency.number_format = "#,##0"
    data_currency.border = THIN_BORDER
    wb.add_named_style(data_currency)

    input_cell = NamedStyle(name="input_cell")
    input_cell.font = Font(name="Calibri", size=10, bold=True)
    input_cell.fill = INPUT_FILL
    input_cell.border = BOLD_BORDER
    wb.add_named_style(input_cell)

    projected = NamedStyle(name="projected")
    projected.font = Font(name="Calibri", size=10)
    projected.fill = PROJECTED_FILL
    projected.border = THIN_BORDER
    wb.add_named_style(projected)


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------


def _auto_col_widths(ws, min_w: int = COL_MIN_WIDTH, max_w: int = COL_MAX_WIDTH) -> None:
    """Set column widths based on content length, clamped to [min_w, max_w]."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        width = min(max(max_len + 2, min_w), max_w)
        ws.column_dimensions[col_letter].width = width


def _freeze_pane(ws, row: int = 2, col: int = 2) -> None:
    """Freeze the first row and first column."""
    ws.freeze_panes = ws.cell(row=row, column=col)


def _write_header_row(ws, row: int, values: list, start_col: int = 1) -> None:
    """Write a row of header-styled cells."""
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.style = "header"


def _is_pct_metric(name: str) -> bool:
    """Return True if a metric name looks like a percentage."""
    lower = name.lower()
    return any(
        kw in lower
        for kw in ["margin", "growth", "yield", "rate", "%", "pct", "percent"]
    )


def _is_eps_metric(name: str) -> bool:
    lower = name.lower()
    return "eps" in lower or "per share" in lower


def _style_for_metric(name: str) -> str:
    if _is_pct_metric(name):
        return "data_pct"
    return "data_currency"


def _number_format_for_metric(name: str) -> str:
    if _is_pct_metric(name):
        return "0.0%"
    if _is_eps_metric(name):
        return "#,##0.00"
    return "#,##0"


def _write_data_cell(ws, row: int, col: int, value, metric_name: str,
                     is_projected: bool = False, is_input: bool = False):
    """Write a single data cell with proper style and number format."""
    cell = ws.cell(row=row, column=col, value=value)
    if is_input:
        cell.style = "input_cell"
        cell.number_format = _number_format_for_metric(metric_name)
    elif is_projected:
        cell.style = "projected"
        cell.number_format = _number_format_for_metric(metric_name)
    elif _is_pct_metric(metric_name):
        cell.style = "data_pct"
    elif _is_eps_metric(metric_name):
        cell.style = "data"
        cell.number_format = "#,##0.00"
    else:
        cell.style = "data_currency"
    return cell


def _add_growth_conditional_formatting(ws, min_row: int, max_row: int,
                                       min_col: int, max_col: int) -> None:
    """Apply green/red conditional formatting for positive/negative values."""
    if min_col > max_col or min_row > max_row:
        return
    cell_range = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThan", formula=["0"], font=GREEN_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT),
    )


def _col_letter(col: int) -> str:
    return get_column_letter(col)


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------


def build_summary(wb: Workbook, ctx: dict) -> None:
    """Tab 1: Summary dashboard with company info and key stats."""
    company = ctx.get("company", {})
    market = ctx.get("market_data", {})
    if not company and not market:
        return

    ws = wb.create_sheet("Summary")

    # Company header
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1,
                         value=f"{company.get('name', '')} ({company.get('ticker', '')})")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1B2A4A")
    title_cell.alignment = Alignment(horizontal="left")

    ws.cell(row=2, column=1, value=f"Exchange: {company.get('exchange', 'N/A')}")
    ws.cell(row=2, column=2, value=f"Currency: {company.get('currency', 'N/A')}")

    # Key stats in 2-column layout
    _write_header_row(ws, 4, ["Metric", "Value"])

    stats = [
        ("Stock Price", market.get("price"), "$#,##0.00"),
        ("Market Cap", market.get("market_cap"), "#,##0"),
        ("Shares Outstanding", market.get("shares_outstanding"), "#,##0"),
        ("Beta", market.get("beta"), "0.00"),
        ("52-Week High", market.get("fifty_two_week_high"), "$#,##0.00"),
        ("52-Week Low", market.get("fifty_two_week_low"), "$#,##0.00"),
        ("Trailing P/E", market.get("trailing_pe"), "0.0x"),
        ("Forward P/E", market.get("forward_pe"), "0.0x"),
        ("EV/EBITDA", market.get("ev_ebitda"), "0.0x"),
        ("Dividend Yield", market.get("dividend_yield"), "0.0%"),
    ]

    for i, (label, value, fmt) in enumerate(stats):
        if value is None:
            continue
        row = 5 + i
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.style = "data"
        label_cell.font = Font(name="Calibri", size=10, bold=True)
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.style = "data"
        val_cell.number_format = fmt

    _auto_col_widths(ws)


def _build_financial_tab(wb: Workbook, ctx: dict, tab_name: str,
                         data_key: str, extra_computed_rows: list | None = None) -> None:
    """Generic builder for Income Statement, Balance Sheet, Cash Flow tabs."""
    data = ctx.get(data_key)
    if not data:
        return

    periods = ctx.get("periods", [])
    projected_periods = ctx.get("projected_periods", [])
    all_periods = periods + projected_periods
    if not all_periods:
        return

    ws = wb.create_sheet(tab_name)

    # Header row
    headers = [""] + all_periods
    _write_header_row(ws, 1, headers)

    # Data rows
    metrics = list(data.keys())
    row_map: dict[str, int] = {}  # metric_name -> row number

    for i, metric in enumerate(metrics):
        row = i + 2
        row_map[metric] = row
        # Label cell
        label_cell = ws.cell(row=row, column=1, value=metric)
        label_cell.style = "data"
        label_cell.font = Font(name="Calibri", size=10, bold=True)

        series = data[metric]
        for j, period in enumerate(all_periods):
            col = j + 2
            value = series.get(period)
            is_proj = period in projected_periods
            _write_data_cell(ws, row, col, value, metric, is_projected=is_proj)

    # Computed rows (margin %, growth %, FCF margin, etc.)
    current_row = len(metrics) + 2

    if extra_computed_rows:
        for computed in extra_computed_rows:
            label = computed["label"]
            numerator_key = computed["numerator"]
            denominator_key = computed["denominator"]

            num_row = row_map.get(numerator_key)
            den_row = row_map.get(denominator_key)
            if num_row is None or den_row is None:
                continue

            row_map[label] = current_row
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True, italic=True)

            for j in range(len(all_periods)):
                col = j + 2
                den_ref = f"{_col_letter(col)}{den_row}"
                num_ref = f"{_col_letter(col)}{num_row}"
                formula = f"=IF({den_ref}<>0,{num_ref}/{den_ref},\"\")"
                cell = ws.cell(row=current_row, column=col, value=formula)
                is_proj = all_periods[j] in projected_periods
                if is_proj:
                    cell.style = "projected"
                else:
                    cell.style = "data_pct"
                cell.number_format = "0.0%"

            current_row += 1

    # YoY growth rows for primary metrics
    growth_metrics = [m for m in metrics if not _is_pct_metric(m)]
    if growth_metrics and len(all_periods) > 4:
        # Blank separator row
        current_row += 1
        for metric in growth_metrics:
            src_row = row_map.get(metric)
            if src_row is None:
                continue
            label = f"{metric} YoY Growth %"
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True, italic=True)

            growth_cols_start = None
            growth_cols_end = None
            for j in range(4, len(all_periods)):
                col = j + 2
                prior_col = j - 4 + 2  # 4 quarters back
                ref_curr = f"{_col_letter(col)}{src_row}"
                ref_prior = f"{_col_letter(prior_col)}{src_row}"
                formula = f'=IF(AND({ref_prior}<>"",{ref_prior}<>0),{ref_curr}/{ref_prior}-1,"")'
                cell = ws.cell(row=current_row, column=col, value=formula)
                is_proj = all_periods[j] in projected_periods
                if is_proj:
                    cell.style = "projected"
                else:
                    cell.style = "data_pct"
                cell.number_format = "0.0%"

                if growth_cols_start is None:
                    growth_cols_start = col
                growth_cols_end = col

            # Conditional formatting on the growth row
            if growth_cols_start and growth_cols_end:
                _add_growth_conditional_formatting(
                    ws, current_row, current_row, growth_cols_start, growth_cols_end
                )
            current_row += 1

    _freeze_pane(ws)
    _auto_col_widths(ws)


def build_income_statement(wb: Workbook, ctx: dict) -> None:
    """Tab 2: Income Statement with computed margin rows."""
    computed = [
        {"label": "Gross Margin %", "numerator": "Gross Profit", "denominator": "Revenue"},
        {"label": "Operating Margin %", "numerator": "Operating Income", "denominator": "Revenue"},
        {"label": "Net Margin %", "numerator": "Net Income", "denominator": "Revenue"},
    ]
    _build_financial_tab(wb, ctx, "Income Statement", "income_statement",
                         extra_computed_rows=computed)


def build_balance_sheet(wb: Workbook, ctx: dict) -> None:
    """Tab 3: Balance Sheet."""
    _build_financial_tab(wb, ctx, "Balance Sheet", "balance_sheet")


def build_cash_flow(wb: Workbook, ctx: dict) -> None:
    """Tab 4: Cash Flow Statement with FCF Margin."""
    computed = [
        {"label": "FCF Margin %", "numerator": "Free Cash Flow", "denominator": "Operating Cash Flow"},
    ]
    _build_financial_tab(wb, ctx, "Cash Flow", "cash_flow",
                         extra_computed_rows=computed)


def build_segments(wb: Workbook, ctx: dict) -> None:
    """Tab 5: Revenue by segment with growth and % of total."""
    segments = ctx.get("segments", {})
    if not segments:
        return

    periods = ctx.get("periods", [])
    projected_periods = ctx.get("projected_periods", [])
    all_periods = periods + projected_periods
    if not all_periods:
        return

    ws = wb.create_sheet("Segments")

    for section_name, section_data in segments.items():
        # Section header
        ws.append([])
        section_start_row = ws.max_row + 1
        ws.cell(row=section_start_row, column=1,
                value=section_name).font = Font(name="Calibri", size=11, bold=True, color="1B2A4A")

        header_row = section_start_row + 1
        _write_header_row(ws, header_row, [""] + all_periods)

        segment_names = list(section_data.keys())
        segment_rows: dict[str, int] = {}
        current_row = header_row + 1

        # Data rows per segment
        for seg_name in segment_names:
            seg_series = section_data[seg_name]
            segment_rows[seg_name] = current_row

            label_cell = ws.cell(row=current_row, column=1, value=seg_name)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True)

            for j, period in enumerate(all_periods):
                col = j + 2
                value = seg_series.get(period)
                is_proj = period in projected_periods
                _write_data_cell(ws, current_row, col, value, seg_name, is_projected=is_proj)

            current_row += 1

        # Total row (sum of segments) using formulas
        total_row = current_row
        label_cell = ws.cell(row=total_row, column=1, value="Total")
        label_cell.style = "data"
        label_cell.font = Font(name="Calibri", size=10, bold=True, underline="single")

        first_seg_row = header_row + 1
        last_seg_row = current_row - 1
        for j in range(len(all_periods)):
            col = j + 2
            formula = f"=SUM({_col_letter(col)}{first_seg_row}:{_col_letter(col)}{last_seg_row})"
            cell = ws.cell(row=total_row, column=col, value=formula)
            is_proj = all_periods[j] in projected_periods
            cell.style = "projected" if is_proj else "data_currency"
            cell.number_format = "#,##0"

        current_row += 1

        # % of Total rows
        current_row += 1  # blank separator
        pct_label = ws.cell(row=current_row, column=1, value="% of Total")
        pct_label.font = Font(name="Calibri", size=10, bold=True, italic=True, color="1B2A4A")
        current_row += 1

        for seg_name in segment_names:
            seg_row = segment_rows[seg_name]
            label_cell = ws.cell(row=current_row, column=1, value=seg_name)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, italic=True)

            for j in range(len(all_periods)):
                col = j + 2
                seg_ref = f"{_col_letter(col)}{seg_row}"
                total_ref = f"{_col_letter(col)}{total_row}"
                formula = f"=IF({total_ref}<>0,{seg_ref}/{total_ref},\"\")"
                cell = ws.cell(row=current_row, column=col, value=formula)
                is_proj = all_periods[j] in projected_periods
                if is_proj:
                    cell.style = "projected"
                else:
                    cell.style = "data_pct"
                cell.number_format = "0.0%"

            current_row += 1

        # YoY Growth rows per segment
        if len(all_periods) > 4:
            current_row += 1  # blank separator
            growth_label = ws.cell(row=current_row, column=1, value="YoY Growth %")
            growth_label.font = Font(name="Calibri", size=10, bold=True, italic=True, color="1B2A4A")
            current_row += 1

            for seg_name in segment_names:
                seg_row = segment_rows[seg_name]
                label_cell = ws.cell(row=current_row, column=1, value=seg_name)
                label_cell.style = "data"
                label_cell.font = Font(name="Calibri", size=10, italic=True)

                growth_col_start = None
                growth_col_end = None
                for j in range(4, len(all_periods)):
                    col = j + 2
                    prior_col = j - 4 + 2
                    ref_curr = f"{_col_letter(col)}{seg_row}"
                    ref_prior = f"{_col_letter(prior_col)}{seg_row}"
                    formula = f'=IF(AND({ref_prior}<>"",{ref_prior}<>0),{ref_curr}/{ref_prior}-1,"")'
                    cell = ws.cell(row=current_row, column=col, value=formula)
                    is_proj = all_periods[j] in projected_periods
                    if is_proj:
                        cell.style = "projected"
                    else:
                        cell.style = "data_pct"
                    cell.number_format = "0.0%"

                    if growth_col_start is None:
                        growth_col_start = col
                    growth_col_end = col

                if growth_col_start and growth_col_end:
                    _add_growth_conditional_formatting(
                        ws, current_row, current_row, growth_col_start, growth_col_end
                    )
                current_row += 1

    _freeze_pane(ws)
    _auto_col_widths(ws)


def build_kpis(wb: Workbook, ctx: dict) -> None:
    """Tab 6: Company-specific operating KPIs."""
    kpis = ctx.get("kpis")
    if not kpis:
        return

    periods = ctx.get("periods", [])
    projected_periods = ctx.get("projected_periods", [])
    all_periods = periods + projected_periods
    if not all_periods:
        return

    ws = wb.create_sheet("KPIs")
    _write_header_row(ws, 1, [""] + all_periods)

    for i, (metric, series) in enumerate(kpis.items()):
        row = i + 2
        label_cell = ws.cell(row=row, column=1, value=metric)
        label_cell.style = "data"
        label_cell.font = Font(name="Calibri", size=10, bold=True)

        for j, period in enumerate(all_periods):
            col = j + 2
            value = series.get(period)
            is_proj = period in projected_periods
            _write_data_cell(ws, row, col, value, metric, is_projected=is_proj)

    _freeze_pane(ws)
    _auto_col_widths(ws)


def build_guidance(wb: Workbook, ctx: dict) -> None:
    """Tab 7: Guidance vs actuals with beat/miss tracking."""
    guidance = ctx.get("guidance")
    if not guidance:
        return

    guide_series = guidance.get("series", {})
    actual_series = guidance.get("actuals", {})
    if not guide_series and not actual_series:
        return

    periods = ctx.get("periods", [])
    projected_periods = ctx.get("projected_periods", [])
    all_periods = periods + projected_periods
    if not all_periods:
        return

    ws = wb.create_sheet("Guidance")
    _write_header_row(ws, 1, [""] + all_periods)

    current_row = 2

    # Build matched pairs: for each guidance series, find its actual counterpart
    guide_keys = list(guide_series.keys())
    actual_keys = list(actual_series.keys())

    # Try to match by replacing "Guidance" with "Actual" or by common prefix
    matched_pairs: list[tuple[str, str | None, str | None]] = []
    used_actuals: set[str] = set()

    for gk in guide_keys:
        base = gk.replace(" Guidance", "").replace(" Guide", "").strip()
        matched_ak = None
        for ak in actual_keys:
            ak_base = ak.replace(" Actual", "").replace(" Act", "").strip()
            if base.lower() == ak_base.lower():
                matched_ak = ak
                used_actuals.add(ak)
                break
        matched_pairs.append((base, gk, matched_ak))

    # Any unmatched actuals
    for ak in actual_keys:
        if ak not in used_actuals:
            base = ak.replace(" Actual", "").replace(" Act", "").strip()
            matched_pairs.append((base, None, ak))

    for base_name, gk, ak in matched_pairs:
        # Section label
        section_cell = ws.cell(row=current_row, column=1, value=base_name)
        section_cell.font = Font(name="Calibri", size=11, bold=True, color="1B2A4A")
        current_row += 1

        guide_row = None
        actual_row = None

        # Guidance row
        if gk:
            guide_row = current_row
            label_cell = ws.cell(row=current_row, column=1, value="Guidance")
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, italic=True)

            g_data = guide_series[gk]
            for j, period in enumerate(all_periods):
                col = j + 2
                value = g_data.get(period)
                _write_data_cell(ws, current_row, col, value, gk)
            current_row += 1

        # Actual row
        if ak:
            actual_row = current_row
            label_cell = ws.cell(row=current_row, column=1, value="Actual")
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, italic=True)

            a_data = actual_series[ak]
            for j, period in enumerate(all_periods):
                col = j + 2
                value = a_data.get(period)
                _write_data_cell(ws, current_row, col, value, ak)
            current_row += 1

        # Beat/Miss row (absolute)
        if guide_row and actual_row:
            bm_row = current_row
            label_cell = ws.cell(row=current_row, column=1, value="Beat / Miss")
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True)

            for j in range(len(all_periods)):
                col = j + 2
                a_ref = f"{_col_letter(col)}{actual_row}"
                g_ref = f"{_col_letter(col)}{guide_row}"
                formula = f'=IF(AND({a_ref}<>"",{g_ref}<>""),{a_ref}-{g_ref},"")'
                cell = ws.cell(row=current_row, column=col, value=formula)
                cell.style = "data"
                cell.number_format = _number_format_for_metric(base_name)

            _add_growth_conditional_formatting(ws, bm_row, bm_row, 2, len(all_periods) + 1)
            current_row += 1

            # Beat/Miss % row
            label_cell = ws.cell(row=current_row, column=1, value="Beat / Miss %")
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True, italic=True)

            for j in range(len(all_periods)):
                col = j + 2
                a_ref = f"{_col_letter(col)}{actual_row}"
                g_ref = f"{_col_letter(col)}{guide_row}"
                formula = (
                    f'=IF(AND({a_ref}<>"",{g_ref}<>"",{g_ref}<>0),'
                    f'{a_ref}/{g_ref}-1,"")'
                )
                cell = ws.cell(row=current_row, column=col, value=formula)
                cell.style = "data_pct"

            _add_growth_conditional_formatting(ws, current_row, current_row, 2, len(all_periods) + 1)
            current_row += 1

        current_row += 1  # blank separator

    _freeze_pane(ws)
    _auto_col_widths(ws)


def build_dcf(wb: Workbook, ctx: dict) -> None:
    """Tab 8: DCF valuation with sensitivity table."""
    dcf = ctx.get("dcf")
    if not dcf:
        return

    ws = wb.create_sheet("DCF")

    # --- Section 1: Key Assumptions ---
    ws.cell(row=1, column=1, value="DCF Valuation").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A"
    )

    _write_header_row(ws, 3, ["Assumption", "Value"])

    assumptions = [
        ("WACC", dcf.get("wacc"), "0.0%"),
        ("Terminal Growth Rate", dcf.get("terminal_growth"), "0.0%"),
        ("Risk-Free Rate", dcf.get("risk_free_rate"), "0.0%"),
        ("Equity Risk Premium", dcf.get("equity_risk_premium"), "0.0%"),
    ]

    row = 4
    for label, value, fmt in assumptions:
        if value is None:
            continue
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.style = "data"
        label_cell.font = Font(name="Calibri", size=10, bold=True)
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.style = "input_cell"
        val_cell.number_format = fmt
        row += 1

    # --- Section 2: FCF Projections ---
    projected_fcf = dcf.get("projected_fcf", [])
    if projected_fcf:
        row += 1
        ws.cell(row=row, column=1, value="Projected Free Cash Flow").font = Font(
            name="Calibri", size=11, bold=True, color="1B2A4A"
        )
        row += 1

        year_headers = [""] + [f"Year {i + 1}" for i in range(len(projected_fcf))]
        _write_header_row(ws, row, year_headers)
        row += 1

        fcf_label = ws.cell(row=row, column=1, value="FCF")
        fcf_label.style = "data"
        fcf_label.font = Font(name="Calibri", size=10, bold=True)
        for i, fcf_val in enumerate(projected_fcf):
            cell = ws.cell(row=row, column=i + 2, value=fcf_val)
            cell.style = "projected"
            cell.number_format = "#,##0"
        row += 1

    # --- Section 3: Valuation Output ---
    row += 1
    ws.cell(row=row, column=1, value="Valuation Output").font = Font(
        name="Calibri", size=11, bold=True, color="1B2A4A"
    )
    row += 1

    outputs = [
        ("Terminal Value", dcf.get("terminal_value"), "#,##0"),
        ("Enterprise Value", dcf.get("enterprise_value"), "#,##0"),
        ("Implied Share Price", dcf.get("implied_share_price"), "$#,##0.00"),
    ]

    for label, value, fmt in outputs:
        if value is None:
            continue
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.style = "data"
        label_cell.font = Font(name="Calibri", size=10, bold=True)
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.style = "data"
        val_cell.number_format = fmt
        row += 1

    # Compare to current price
    market = ctx.get("market_data", {})
    current_price = market.get("price")
    implied = dcf.get("implied_share_price")
    if current_price and implied:
        row += 1
        ws.cell(row=row, column=1, value="Current Price").font = Font(
            name="Calibri", size=10, bold=True
        )
        ws.cell(row=row, column=2, value=current_price).number_format = "$#,##0.00"
        row += 1
        ws.cell(row=row, column=1, value="Upside / Downside").font = Font(
            name="Calibri", size=10, bold=True
        )
        upside_cell = ws.cell(row=row, column=2,
                              value=(implied - current_price) / current_price)
        upside_cell.number_format = "0.0%"
        if implied > current_price:
            upside_cell.font = GREEN_FONT
        else:
            upside_cell.font = RED_FONT

    # --- Section 4: Sensitivity Table ---
    sensitivity = dcf.get("sensitivity")
    if sensitivity:
        wacc_vals = sensitivity.get("wacc_values", [])
        growth_vals = sensitivity.get("growth_values", [])
        prices = sensitivity.get("prices", [])

        if wacc_vals and growth_vals and prices:
            row += 2
            ws.cell(row=row, column=1, value="Sensitivity: Implied Share Price").font = Font(
                name="Calibri", size=11, bold=True, color="1B2A4A"
            )
            row += 1

            # Corner label
            corner_cell = ws.cell(row=row, column=1, value="WACC \\ Terminal Growth")
            corner_cell.style = "header"

            # Column headers = terminal growth rates
            for j, g in enumerate(growth_vals):
                cell = ws.cell(row=row, column=j + 2, value=g / 100.0)
                cell.style = "header"
                cell.number_format = "0.0%"

            row += 1

            # Rows = WACC values
            for i, w in enumerate(wacc_vals):
                wacc_cell = ws.cell(row=row, column=1, value=w / 100.0)
                wacc_cell.style = "data"
                wacc_cell.font = Font(name="Calibri", size=10, bold=True)
                wacc_cell.number_format = "0.0%"

                if i < len(prices):
                    price_row = prices[i]
                    for j, p in enumerate(price_row):
                        cell = ws.cell(row=row, column=j + 2, value=p)
                        cell.style = "data"
                        cell.number_format = "$#,##0.00"

                        # Highlight the cell nearest to current assumptions
                        target_wacc = dcf.get("wacc", 0) * 100
                        target_growth = dcf.get("terminal_growth", 0) * 100
                        if (abs(w - target_wacc) < 0.01
                                and j < len(growth_vals)
                                and abs(growth_vals[j] - target_growth) < 0.01):
                            cell.fill = INPUT_FILL
                            cell.font = Font(name="Calibri", size=10, bold=True)

                row += 1

    _auto_col_widths(ws)


def build_comps(wb: Workbook, ctx: dict) -> None:
    """Tab 9: Comparable company analysis."""
    comps = ctx.get("comps")
    if not comps:
        return

    peers = comps.get("peers", [])
    if not peers:
        return

    ws = wb.create_sheet("Comps")

    # Determine columns from peer data
    # Standard comp columns
    comp_fields = [
        ("Ticker", "ticker", None),
        ("Name", "name", None),
        ("Trailing P/E", "trailing_pe", "0.0x"),
        ("EV/EBITDA", "ev_ebitda", "0.0x"),
        ("Price/Sales", "price_to_sales", "0.0x"),
        ("Revenue Growth", "revenue_growth", "0.0%"),
        ("Operating Margin", "op_margin", "0.0%"),
    ]

    # Filter to fields that exist in at least one peer
    active_fields: list[tuple[str, str, str | None]] = []
    for label, key, fmt in comp_fields:
        if any(key in p for p in peers):
            active_fields.append((label, key, fmt))

    if not active_fields:
        return

    # Header
    _write_header_row(ws, 1, [f[0] for f in active_fields])

    # Target company row
    company = ctx.get("company", {})
    market = ctx.get("market_data", {})
    target_ticker = company.get("ticker", "Target")

    row = 2
    target_cell = ws.cell(row=row, column=1, value=target_ticker)
    target_cell.style = "data"
    target_cell.font = Font(name="Calibri", size=10, bold=True)
    target_cell.fill = INPUT_FILL

    target_map = {
        "ticker": target_ticker,
        "name": company.get("name", ""),
        "trailing_pe": market.get("trailing_pe"),
        "ev_ebitda": market.get("ev_ebitda"),
    }
    for j, (label, key, fmt) in enumerate(active_fields):
        if j == 0:
            continue  # already wrote ticker
        value = target_map.get(key)
        if value is not None:
            cell = ws.cell(row=row, column=j + 1, value=value)
            cell.style = "data"
            cell.fill = INPUT_FILL
            if fmt:
                cell.number_format = fmt

    row += 1

    # Peer rows
    peer_start_row = row
    for peer in peers:
        for j, (label, key, fmt) in enumerate(active_fields):
            value = peer.get(key)
            cell = ws.cell(row=row, column=j + 1, value=value)
            cell.style = "data"
            if fmt:
                cell.number_format = fmt
        row += 1
    peer_end_row = row - 1

    # Blank row
    row += 1

    # Median row
    median_row = row
    med_label = ws.cell(row=row, column=1, value="Median")
    med_label.style = "data"
    med_label.font = Font(name="Calibri", size=10, bold=True, underline="single")

    for j, (label, key, fmt) in enumerate(active_fields):
        if j < 2:  # skip ticker and name
            continue
        col = j + 1
        col_l = _col_letter(col)
        formula = f"=MEDIAN({col_l}{peer_start_row}:{col_l}{peer_end_row})"
        cell = ws.cell(row=row, column=col, value=formula)
        cell.style = "data"
        cell.font = Font(name="Calibri", size=10, bold=True)
        if fmt:
            cell.number_format = fmt
    row += 1

    # Mean row
    mean_label = ws.cell(row=row, column=1, value="Mean")
    mean_label.style = "data"
    mean_label.font = Font(name="Calibri", size=10, bold=True, underline="single")

    for j, (label, key, fmt) in enumerate(active_fields):
        if j < 2:
            continue
        col = j + 1
        col_l = _col_letter(col)
        formula = f"=AVERAGE({col_l}{peer_start_row}:{col_l}{peer_end_row})"
        cell = ws.cell(row=row, column=col, value=formula)
        cell.style = "data"
        cell.font = Font(name="Calibri", size=10, bold=True)
        if fmt:
            cell.number_format = fmt
    row += 1

    # Implied Value row (apply peer median multiples to target fundamentals)
    row += 1
    iv_label = ws.cell(row=row, column=1, value="Implied Value")
    iv_label.style = "data"
    iv_label.font = Font(name="Calibri", size=10, bold=True, color="1B2A4A")
    iv_label.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # For P/E: implied price = Median P/E * Target EPS
    # For EV/EBITDA: implied EV = Median EV/EBITDA * Target EBITDA
    # We note these conceptually; actual computation requires EPS/EBITDA data
    income = ctx.get("income_statement", {})
    eps_data = income.get("EPS", {})
    periods = ctx.get("periods", [])
    latest_eps = None
    if eps_data and periods:
        for p in reversed(periods):
            if p in eps_data and eps_data[p] is not None:
                latest_eps = eps_data[p]
                break

    for j, (label, key, fmt) in enumerate(active_fields):
        if j < 2:
            continue
        col = j + 1
        if key == "trailing_pe" and latest_eps is not None:
            # Implied price = Median P/E * trailing EPS
            col_l = _col_letter(col)
            formula = f"={col_l}{median_row}*{latest_eps}"
            cell = ws.cell(row=row, column=col, value=formula)
            cell.style = "data"
            cell.number_format = "$#,##0.00"
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        else:
            # For other multiples, leave a note
            cell = ws.cell(row=row, column=col, value="--")
            cell.style = "data"
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    _freeze_pane(ws, row=2, col=3)
    _auto_col_widths(ws)


def build_projections(wb: Workbook, ctx: dict) -> None:
    """Tab 10: Projections with yellow input assumption cells."""
    projections = ctx.get("projections")
    assumptions = ctx.get("projection_assumptions")
    if not projections and not assumptions:
        return

    projected_periods = ctx.get("projected_periods", [])
    if not projected_periods:
        return

    ws = wb.create_sheet("Projections")

    # --- Section 1: Assumptions (input cells) ---
    ws.cell(row=1, column=1, value="Projection Assumptions").font = Font(
        name="Calibri", size=14, bold=True, color="1B2A4A"
    )
    ws.cell(row=2, column=1, value="Yellow cells are analyst inputs").font = Font(
        name="Calibri", size=9, italic=True, color="666666"
    )

    current_row = 4

    if assumptions:
        _write_header_row(ws, current_row, ["Assumption"] + projected_periods)
        current_row += 1

        # Time-varying assumptions
        time_varying_keys = [
            ("revenue_growth", "Revenue Growth"),
            ("gross_margin", "Gross Margin"),
            ("op_margin", "Operating Margin"),
            ("capex_pct_revenue", "CapEx % of Revenue"),
        ]

        for key, label in time_varying_keys:
            series = assumptions.get(key)
            if series is None:
                continue
            if isinstance(series, dict):
                label_cell = ws.cell(row=current_row, column=1, value=label)
                label_cell.style = "data"
                label_cell.font = Font(name="Calibri", size=10, bold=True)

                for j, period in enumerate(projected_periods):
                    col = j + 2
                    value = series.get(period)
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.style = "input_cell"
                    cell.number_format = "0.0%"

                current_row += 1

        # Scalar assumptions
        scalar_keys = [
            ("tax_rate", "Tax Rate", "0.0%"),
            ("buyback_rate_qoq", "Buyback Rate (QoQ)", "0.0%"),
        ]

        for key, label, fmt in scalar_keys:
            value = assumptions.get(key)
            if value is None:
                continue
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True)

            val_cell = ws.cell(row=current_row, column=2, value=value)
            val_cell.style = "input_cell"
            val_cell.number_format = fmt

            # Merge label across remaining columns for clarity
            current_row += 1

    # --- Section 2: Projected Financials ---
    if projections:
        current_row += 2
        ws.cell(row=current_row, column=1, value="Projected Financials").font = Font(
            name="Calibri", size=11, bold=True, color="1B2A4A"
        )
        current_row += 1

        _write_header_row(ws, current_row, [""] + projected_periods)
        current_row += 1

        for metric, series in projections.items():
            label_cell = ws.cell(row=current_row, column=1, value=metric)
            label_cell.style = "data"
            label_cell.font = Font(name="Calibri", size=10, bold=True)

            for j, period in enumerate(projected_periods):
                col = j + 2
                value = series.get(period)
                _write_data_cell(ws, current_row, col, value, metric, is_projected=True)

            current_row += 1

    _freeze_pane(ws)
    _auto_col_widths(ws)


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

TAB_BUILDERS = [
    ("company", build_summary),
    ("income_statement", build_income_statement),
    ("balance_sheet", build_balance_sheet),
    ("cash_flow", build_cash_flow),
    ("segments", build_segments),
    ("kpis", build_kpis),
    ("guidance", build_guidance),
    ("dcf", build_dcf),
    ("comps", build_comps),
    ("projections", build_projections),
]

# Summary tab triggers on either company or market_data
_SUMMARY_KEYS = {"company", "market_data"}


def _merge_projections_into_financials(ctx: dict) -> None:
    """Merge projected values into income_statement, balance_sheet, and cash_flow.

    For each metric in ctx["projections"], if a matching metric name exists in
    one of the financial statement sections, the projected period values are
    added to that metric's dict so they appear inline in the financial tabs.
    """
    projections = ctx.get("projections")
    if not projections:
        return
    target_sections = ["income_statement", "balance_sheet", "cash_flow"]
    for metric, proj_values in projections.items():
        for section_key in target_sections:
            section = ctx.get(section_key)
            if section and metric in section:
                section[metric].update(proj_values)
                break


def _historical_ratio(section: dict, numerator: str, denominator: str,
                      periods: list) -> float | None:
    """Compute trailing average ratio of numerator/denominator over recent periods."""
    num_series = section.get(numerator, {})
    den_series = section.get(denominator, {})
    ratios = []
    for p in reversed(periods):
        n = num_series.get(p)
        d = den_series.get(p)
        if n is not None and d is not None and d != 0:
            ratios.append(n / d)
        if len(ratios) >= 4:
            break
    return sum(ratios) / len(ratios) if ratios else None


def _trailing_avg(section: dict, metric: str, periods: list, n: int = 4) -> float | None:
    """Compute trailing average of a metric over the last n available periods."""
    series = section.get(metric, {})
    vals = []
    for p in reversed(periods):
        v = series.get(p)
        if v is not None:
            vals.append(v)
        if len(vals) >= n:
            break
    return sum(vals) / len(vals) if vals else None


def _last_value(section: dict, metric: str, periods: list):
    """Get the most recent non-None value for a metric."""
    series = section.get(metric, {})
    for p in reversed(periods):
        v = series.get(p)
        if v is not None:
            return v
    return None


def _enrich_projections(ctx: dict) -> None:
    """Derive sub-line items for projected periods from high-level projections.

    Uses historical ratios to fill in IS/CF/BS items that weren't directly
    projected. Runs after _merge_projections_into_financials().
    """
    hist_periods = ctx.get("periods", [])
    proj_periods = ctx.get("projected_periods", [])
    if not hist_periods or not proj_periods:
        return

    is_data = ctx.get("income_statement", {})
    cf_data = ctx.get("cash_flow", {})
    bs_data = ctx.get("balance_sheet", {})
    assumptions = ctx.get("projection_assumptions", {})
    tax_rate = assumptions.get("tax_rate", 0.16)

    # ---- Income Statement derivations ----
    for p in proj_periods:
        rev = is_data.get("Revenue", {}).get(p)
        gp = is_data.get("Gross Profit", {}).get(p)
        op_inc = is_data.get("Operating Income", {}).get(p)
        ni = is_data.get("Net Income", {}).get(p)
        if rev is None:
            continue

        # Cost of Sales = Revenue - Gross Profit
        if gp is not None:
            is_data.setdefault("Cost of Sales", {})[p] = round(rev - gp)

        # D&A: project as % of revenue
        da_ratio = _historical_ratio(is_data, "D&A", "Revenue", hist_periods)
        if da_ratio is not None:
            is_data.setdefault("D&A", {})[p] = round(rev * da_ratio)

        # R&D: project as % of revenue
        rd_ratio = _historical_ratio(is_data, "Research & Development", "Revenue", hist_periods)
        if rd_ratio is not None:
            is_data.setdefault("Research & Development", {})[p] = round(rev * rd_ratio)

        # SG&A: project as % of revenue
        sga_ratio = _historical_ratio(is_data, "Selling, General & Administrative", "Revenue", hist_periods)
        if sga_ratio is not None:
            is_data.setdefault("Selling, General & Administrative", {})[p] = round(rev * sga_ratio)

        # Total Operating Expenses = Revenue - Operating Income
        if op_inc is not None:
            is_data.setdefault("Total Operating Expenses", {})[p] = round(rev - op_inc)

        # Other Income/(Expense): trailing average
        other_avg = _trailing_avg(is_data, "Other Income/(Expense)", hist_periods)
        if other_avg is not None:
            is_data.setdefault("Other Income/(Expense)", {})[p] = round(other_avg)

        # Pre-tax Income = Operating Income + Other Income
        other_inc = is_data.get("Other Income/(Expense)", {}).get(p, 0)
        if op_inc is not None:
            pretax = round(op_inc + (other_inc or 0))
            is_data.setdefault("Pre-tax Income", {})[p] = pretax

            # Tax Provision = Pre-tax - Net Income, or Pre-tax × tax_rate
            if ni is not None:
                is_data.setdefault("Tax Provision", {})[p] = round(pretax - ni)
            else:
                is_data.setdefault("Tax Provision", {})[p] = round(pretax * tax_rate)

    # ---- Cash Flow derivations ----
    for p in proj_periods:
        rev = is_data.get("Revenue", {}).get(p)
        if rev is None:
            continue

        # D&A (same as IS)
        da_val = is_data.get("D&A", {}).get(p)
        if da_val is not None:
            cf_data.setdefault("Depreciation & Amortization", {})[p] = da_val

        # SBC: project as % of revenue
        sbc_ratio = _historical_ratio(cf_data, "Share-based Compensation", "Operating Cash Flow", hist_periods)
        if sbc_ratio is not None:
            ocf = cf_data.get("Operating Cash Flow", {}).get(p)
            if ocf is not None:
                cf_data.setdefault("Share-based Compensation", {})[p] = round(ocf * sbc_ratio)

        # Dividends Paid: trailing average
        div_avg = _trailing_avg(cf_data, "Dividends Paid", hist_periods)
        if div_avg is not None:
            cf_data.setdefault("Dividends Paid", {})[p] = round(div_avg)

        # Share Repurchases: trailing average
        buyback_avg = _trailing_avg(cf_data, "Share Repurchases", hist_periods)
        if buyback_avg is not None:
            cf_data.setdefault("Share Repurchases", {})[p] = round(buyback_avg)

        # Net Cash from Investing ≈ CapEx (simplified)
        capex = cf_data.get("Capital Expenditures", {}).get(p)
        if capex is not None:
            cf_data.setdefault("Net Cash from Investing", {})[p] = round(capex)

        # Net Cash from Financing ≈ Dividends + Buybacks + Debt changes
        div = cf_data.get("Dividends Paid", {}).get(p, 0)
        buybacks = cf_data.get("Share Repurchases", {}).get(p, 0)
        cf_data.setdefault("Net Cash from Financing", {})[p] = round((div or 0) + (buybacks or 0))

    # ---- Balance Sheet derivations ----
    all_periods = hist_periods + proj_periods
    for i, p in enumerate(proj_periods):
        # Find the prior period
        idx_in_all = all_periods.index(p)
        if idx_in_all == 0:
            continue
        prior_p = all_periods[idx_in_all - 1]

        rev = is_data.get("Revenue", {}).get(p)
        if rev is None:
            continue

        # PP&E = prior PP&E + CapEx(abs) - D&A
        prior_ppe = bs_data.get("PP&E (net)", {}).get(prior_p)
        capex = cf_data.get("Capital Expenditures", {}).get(p)
        da = is_data.get("D&A", {}).get(p)
        if prior_ppe is not None and capex is not None and da is not None:
            # CapEx is stored negative; adding abs value
            bs_data.setdefault("PP&E (net)", {})[p] = round(prior_ppe + abs(capex) - da)

        # Current assets/liabilities: project as % of revenue (DSO/DPO approach)
        for bs_metric in ["Accounts Receivable", "Inventories", "Accounts Payable",
                          "Deferred Revenue (Current)", "Other Current Assets",
                          "Other Current Liabilities"]:
            ratio = _historical_ratio(bs_data, bs_metric, is_data.get("Revenue") and "Revenue",
                                      hist_periods)
            # Use direct ratio to revenue from historical
            hist_series = bs_data.get(bs_metric, {})
            rev_series = is_data.get("Revenue", {})
            ratios = []
            for hp in reversed(hist_periods):
                bv = hist_series.get(hp)
                rv = rev_series.get(hp)
                if bv is not None and rv is not None and rv != 0:
                    ratios.append(bv / rv)
                if len(ratios) >= 4:
                    break
            if ratios:
                avg_ratio = sum(ratios) / len(ratios)
                bs_data.setdefault(bs_metric, {})[p] = round(rev * avg_ratio)

        # Short/Long-term Investments: hold flat at last known value
        for flat_metric in ["Short-term Investments", "Long-term Investments",
                            "Vendor Non-trade Receivables"]:
            if flat_metric not in bs_data:
                continue
            if bs_data[flat_metric].get(p) is None:
                last = _last_value(bs_data, flat_metric, hist_periods + proj_periods[:proj_periods.index(p)])
                if last is not None:
                    bs_data[flat_metric][p] = last

        # Debt: hold flat
        for debt_metric in ["Commercial Paper", "Current Term Debt", "Long-term Debt"]:
            if debt_metric not in bs_data:
                continue
            if bs_data[debt_metric].get(p) is None:
                last = _last_value(bs_data, debt_metric, hist_periods + proj_periods[:proj_periods.index(p)])
                if last is not None:
                    bs_data[debt_metric][p] = last

        # Other Non-current Assets/Liabilities: hold flat
        for flat_metric in ["Other Non-current Assets", "Other Non-current Liabilities"]:
            if flat_metric not in bs_data:
                continue
            if bs_data[flat_metric].get(p) is None:
                last = _last_value(bs_data, flat_metric, hist_periods + proj_periods[:proj_periods.index(p)])
                if last is not None:
                    bs_data[flat_metric][p] = last

        # Compute totals
        # Total Current Assets
        ca_items = ["Cash & Equivalents", "Short-term Investments", "Accounts Receivable",
                     "Inventories", "Vendor Non-trade Receivables", "Other Current Assets"]
        ca_vals = [bs_data.get(m, {}).get(p) for m in ca_items]
        if any(v is not None for v in ca_vals):
            bs_data.setdefault("Total Current Assets", {})[p] = round(sum(v for v in ca_vals if v is not None))

        # Total Non-current Assets
        nca_items = ["Long-term Investments", "PP&E (net)", "Other Non-current Assets"]
        nca_vals = [bs_data.get(m, {}).get(p) for m in nca_items]
        if any(v is not None for v in nca_vals):
            bs_data.setdefault("Total Non-current Assets", {})[p] = round(sum(v for v in nca_vals if v is not None))

        # Total Assets
        tca = bs_data.get("Total Current Assets", {}).get(p)
        tnca = bs_data.get("Total Non-current Assets", {}).get(p)
        if tca is not None and tnca is not None:
            bs_data.setdefault("Total Assets", {})[p] = tca + tnca

        # Total Current Liabilities
        cl_items = ["Accounts Payable", "Deferred Revenue (Current)", "Commercial Paper",
                     "Current Term Debt", "Other Current Liabilities"]
        cl_vals = [bs_data.get(m, {}).get(p) for m in cl_items]
        if any(v is not None for v in cl_vals):
            bs_data.setdefault("Total Current Liabilities", {})[p] = round(sum(v for v in cl_vals if v is not None))

        # Total Non-current Liabilities
        ncl_items = ["Long-term Debt", "Other Non-current Liabilities"]
        ncl_vals = [bs_data.get(m, {}).get(p) for m in ncl_items]
        if any(v is not None for v in ncl_vals):
            bs_data.setdefault("Total Non-current Liabilities", {})[p] = round(sum(v for v in ncl_vals if v is not None))

        # Total Liabilities
        tcl = bs_data.get("Total Current Liabilities", {}).get(p)
        tncl = bs_data.get("Total Non-current Liabilities", {}).get(p)
        if tcl is not None and tncl is not None:
            bs_data.setdefault("Total Liabilities", {})[p] = tcl + tncl

        # Cash = prior cash + FCF - dividends(abs) - buybacks(abs)
        prior_cash = bs_data.get("Cash & Equivalents", {}).get(prior_p)
        fcf = cf_data.get("Free Cash Flow", {}).get(p)
        div = cf_data.get("Dividends Paid", {}).get(p, 0)
        buybacks = cf_data.get("Share Repurchases", {}).get(p, 0)
        if prior_cash is not None and fcf is not None:
            # Dividends and buybacks are stored as negative (outflows)
            bs_data.setdefault("Cash & Equivalents", {})[p] = round(
                prior_cash + fcf + (div or 0) + (buybacks or 0)
            )

        # Shareholders' Equity = prior equity + net income + dividends + buybacks
        prior_eq = bs_data.get("Total Shareholders Equity", {}).get(prior_p)
        ni = is_data.get("Net Income", {}).get(p)
        if prior_eq is not None and ni is not None:
            bs_data.setdefault("Total Shareholders Equity", {})[p] = round(
                prior_eq + ni + (div or 0) + (buybacks or 0)
            )


def build_workbook(ctx: dict) -> Workbook:
    """Build the full multi-tab workbook from context data."""
    _merge_projections_into_financials(ctx)
    _enrich_projections(ctx)

    wb = Workbook()
    register_styles(wb)

    # Remove the default sheet created by openpyxl
    default_sheet = wb.active
    tabs_created = 0

    for trigger_key, builder_fn in TAB_BUILDERS:
        # Summary triggers on company OR market_data
        if trigger_key == "company":
            if ctx.get("company") or ctx.get("market_data"):
                builder_fn(wb, ctx)
                tabs_created += 1
        # Projections triggers on projections OR projection_assumptions
        elif trigger_key == "projections":
            if ctx.get("projections") or ctx.get("projection_assumptions"):
                builder_fn(wb, ctx)
                tabs_created += 1
        else:
            if ctx.get(trigger_key):
                builder_fn(wb, ctx)
                tabs_created += 1

    # Remove the default empty sheet if we created any tabs
    if tabs_created > 0 and default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a multi-tab financial model Excel file from context JSON."
    )
    parser.add_argument(
        "--context", required=True,
        help="Path to the input context JSON file."
    )
    parser.add_argument(
        "--output", required=True,
        help="Path for the output .xlsx file."
    )
    args = parser.parse_args()

    # Load context
    context_path = Path(args.context)
    if not context_path.is_file():
        print(f"Error: context file not found: {args.context}", file=sys.stderr)
        sys.exit(1)

    try:
        with open(context_path, "r") as f:
            ctx = json.load(f)
    except json.JSONDecodeError as e:
        print(f"Error: invalid JSON in context file: {e}", file=sys.stderr)
        sys.exit(1)

    # Build workbook
    try:
        wb = build_workbook(ctx)
    except Exception as e:
        print(f"Error building workbook: {e}", file=sys.stderr)
        sys.exit(1)

    # Save
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(output_path))
    except Exception as e:
        print(f"Error saving workbook: {e}", file=sys.stderr)
        sys.exit(1)

    print(str(output_path))


if __name__ == "__main__":
    main()
